#!/usr/bin/env python3
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl，嘗試自動安裝...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--user", "openpyxl"])
    import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
INPUT_DIR = SCRIPT_DIR / "input"
OUTPUT_DIR = SCRIPT_DIR / "output"

EXPECTED_HEADERS = ["題目", "選項一", "選項二", "選項三", "選項四", "正確答案", "解析"]

DIFFICULTY_CHOICES = ["easy", "medium", "hard"]
STYLE_CHOICES = ["情境題", "反面題", "比較題", "綜合題", ""]
QUESTION_TYPE_CHOICES = ["single", "choice"]


def ask(prompt, default=None, choices=None, allow_empty=False):
    hint = ""
    if choices:
        hint = f" [{' / '.join(c if c else '(空白)' for c in choices)}]"
    if default is not None:
        hint += f" (預設: {default!r})"
    while True:
        raw = input(f"{prompt}{hint}: ").strip()
        if not raw:
            if default is not None:
                return default
            if allow_empty:
                return ""
            print("  ⚠️  此欄位不可空白，請重新輸入")
            continue
        if choices is not None and raw not in choices:
            print(f"  ⚠️  只能輸入: {choices}")
            continue
        return raw


def ask_int(prompt, default=None):
    while True:
        raw = input(f"{prompt}" + (f" (預設: {default})" if default is not None else "") + ": ").strip()
        if not raw and default is not None:
            return default
        try:
            return int(raw)
        except ValueError:
            print("  ⚠️  請輸入整數")


def guess_difficulty(filename):
    name = filename.lower()
    if "初級" in filename or "easy" in name:
        return "easy"
    if "中級" in filename or "medium" in name:
        return "medium"
    if "高級" in filename or "進階" in filename or "hard" in name:
        return "hard"
    return None


def read_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError(f"{path.name}: 檔案沒有資料列")
    header = rows[0]
    if len(header) < 6:
        raise ValueError(f"{path.name}: 欄位數不足 6 欄，實際 {len(header)} 欄")
    data_rows = []
    for idx, row in enumerate(rows[1:], start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        data_rows.append((idx, row))
    return data_rows


def convert_file(path):
    print(f"\n{'=' * 60}")
    print(f"📄 處理檔案: {path.name}")
    print(f"{'=' * 60}")

    data_rows = read_excel(path)
    print(f"共讀到 {len(data_rows)} 筆資料列\n")

    print("請為這個檔案設定缺失的欄位（套用到本檔所有題目）：\n")
    subject_no = ask_int("科目編號 (subject_no)", default=1)
    subject = ask("科目名稱 (subject)")
    chapter_no = ask_int("章節編號 (chapter_no)", default=1)
    chapter = ask("章節名稱 (chapter)")
    question_type = ask("題型 (question_type)", default="single", choices=QUESTION_TYPE_CHOICES)

    guessed = guess_difficulty(path.stem)
    difficulty = ask("難度 (difficulty)", default=guessed or "medium", choices=DIFFICULTY_CHOICES)

    question_style = ask(
        "題型風格 (question_style，可留空)",
        default="",
        choices=STYLE_CHOICES,
        allow_empty=True,
    )

    questions = []
    skipped = []
    for seq, (excel_row_no, row) in enumerate(data_rows, start=1):
        cells = [("" if v is None else str(v).strip()) for v in row]
        while len(cells) < 7:
            cells.append("")

        question_text, opt_a, opt_b, opt_c, opt_d, correct, explanation = cells[:7]

        if not question_text:
            skipped.append((excel_row_no, "題目空白"))
            continue

        item = {
            "subject_no": subject_no,
            "subject": subject,
            "chapter_no": chapter_no,
            "chapter": chapter,
            "question_no": seq,
            "question": question_text,
            "option_a": opt_a,
            "option_b": opt_b,
            "option_c": opt_c,
            "option_d": opt_d,
            "correct_answer": correct,
            "question_type": question_type,
            "explanation": explanation,
            "difficulty": difficulty,
        }
        if question_style:
            item["question_style"] = question_style

        questions.append(item)

    return questions, skipped


def safe_output_name(stem):
    cleaned = re.sub(r"[^\w一-鿿-]+", "_", stem).strip("_")
    return cleaned or "output"


def move_source_to_output(src_path, timestamp):
    """把已成功轉檔的 Excel 從 input/ 搬到 output/，撞名時自動加時間戳。"""
    target = OUTPUT_DIR / src_path.name
    if target.exists():
        target = OUTPUT_DIR / f"{src_path.stem}_{timestamp}{src_path.suffix}"
    try:
        shutil.move(str(src_path), str(target))
        return target.name
    except Exception as e:
        print(f"  ⚠️  搬移 {src_path.name} 失敗: {e}")
        return None


def main():
    print("🎯 Excel → 題庫 JSON 轉換器")
    print(f"輸入資料夾: {INPUT_DIR}")
    print(f"輸出資料夾: {OUTPUT_DIR}")

    xlsx_files = sorted(INPUT_DIR.glob("*.xlsx"))
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

    if not xlsx_files:
        print(f"\n⚠️  {INPUT_DIR} 裡沒有 .xlsx 檔案")
        print("請把 Excel 題庫檔放進去，然後再次雙擊 convert.command")
        return

    print(f"\n找到 {len(xlsx_files)} 個檔案：")
    for f in xlsx_files:
        print(f"  - {f.name}")

    all_results = []
    for path in xlsx_files:
        try:
            questions, skipped = convert_file(path)
        except Exception as e:
            print(f"\n❌ {path.name} 轉換失敗: {e}")
            all_results.append((path, None, [], str(e), None))
            continue

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_name = f"{safe_output_name(path.stem)}_{timestamp}.json"
        output_path = OUTPUT_DIR / output_name
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(questions, f, ensure_ascii=False, indent=2)

        print(f"\n✅ 已產出 {len(questions)} 題 → {output_path.name}")
        if skipped:
            print(f"   跳過 {len(skipped)} 列：")
            for row_no, reason in skipped:
                print(f"     - 第 {row_no} 列：{reason}")

        moved_name = move_source_to_output(path, timestamp)
        if moved_name:
            print(f"📦 已將原始 Excel 搬移到 output/{moved_name}")

        all_results.append((path, output_path, skipped, None, moved_name))

    print(f"\n{'=' * 60}")
    print("🎉 全部完成")
    print(f"{'=' * 60}")
    for path, output_path, skipped, err, moved_name in all_results:
        if err:
            print(f"  ❌ {path.name}: {err}")
        else:
            suffix = f"，Excel 已搬到 output/{moved_name}" if moved_name else "，原始 Excel 搬移失敗（仍在 input/）"
            print(f"  ✅ {path.name} → output/{output_path.name}（跳過 {len(skipped)} 列{suffix}）")

    print(f"\n輸出檔在：{OUTPUT_DIR}")
    print("可直接開啟，將 JSON 內容貼到 import.html 匯入。")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n中斷離開。")
    except Exception as e:
        print(f"\n❌ 發生錯誤: {e}")
        import traceback
        traceback.print_exc()
    finally:
        input("\n按 Enter 關閉視窗...")
